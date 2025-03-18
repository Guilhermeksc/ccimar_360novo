from PyQt6.QtCore import QObject
from PyQt6.QtWidgets import QFileDialog, QMessageBox, QInputDialog
from PyQt6.QtCore import Qt
import pandas as pd
import chardet
import webbrowser
from .dashboard.dash_popup import DashboardPopup
import zipfile
import os
import sys
import subprocess
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import chardet
import zipfile
import requests
from io import BytesIO
from datetime import datetime

class CartaoCorporativoController(QObject): 
    def __init__(self, icons, view, model):
        super().__init__()
        self.icons = icons
        self.view = view
        self.model = model  # 🔹 Agora estamos passando diretamente o QSqlTableModel

        self.setup_connections()

    def setup_connections(self):
        """Conecta os sinais da View ao Controller"""
        self.view.refreshRequested.connect(self.import_zip_to_db)
        self.view.rowDoubleClicked.connect(self.row_double_clicked)
        self.view.linkDataCartaoPagamentoGov.connect(self.open_link)
        self.view.exportTable.connect(self.export_table)
        self.view.open_dashboard.connect(self.open_dashboard)  # 🔹 Conecta o botão ao métod
        
    def open_dashboard(self):
        """Abre o popup do dashboard com os dados do model."""
        data = self.model.get_all_model_data()

        if not data:
            QMessageBox.warning(self.view, "Aviso", "Nenhum dado disponível para exibir no Dashboard.")
            return

        # 🔹 Converter para DataFrame caso `data` seja uma lista de dicionários
        df_unique_orgaos = pd.DataFrame(data)[["cod_orgao", "nome_orgao"]].drop_duplicates()

        self.dashboard_popup = DashboardPopup(df_unique_orgaos, self.model.get_data_for_orgao)

        self.dashboard_popup.exec()

    def open_link(self):
        """Abre o link do Portal da Transparência no navegador."""
        url = "https://portaldatransparencia.gov.br/download-de-dados/cpgf"
        webbrowser.open(url)  # 🔹 Abre o link no navegador padrão
    
    def export_table(self):
        """Export all table data, including hidden columns, to an Excel file."""
        file_dialog = QFileDialog()
        file_path, _ = file_dialog.getSaveFileName(
            self.view, "Save as Excel", "", "Excel Files (*.xlsx)"
        )

        if not file_path:
            return  # If the user cancels, do nothing

        # Ensure the file has the correct extension
        if not file_path.lower().endswith(".xlsx"):
            file_path += ".xlsx"

        try:
            headers = []
            data = []

            # Get column headers, including hidden columns
            for col in range(self.view.model.columnCount()):
                header_text = self.view.model.headerData(col, Qt.Orientation.Horizontal)
                headers.append(header_text)

            # Get all data from the model, including hidden columns
            for row in range(self.view.proxy_model.rowCount()):
                row_data = []
                source_index = self.view.proxy_model.mapToSource(self.view.proxy_model.index(row, 0)).row()
                for col in range(self.view.model.columnCount()):
                    index = self.view.model.index(source_index, col)
                    row_data.append(index.data(Qt.ItemDataRole.DisplayRole))
                data.append(row_data)

            # Create DataFrame and export
            df = pd.DataFrame(data, columns=headers)
            df.to_excel(file_path, index=False)

            # Adjust column widths
            column_widths = {
                0: 50, 1: 50, 2: 50, 3: 50, 4: 70, 5: 100, 6: 70, 7: 100
            }
            default_width = 80

            wb = load_workbook(file_path)
            ws = wb.active

            for col_num, column_cells in enumerate(ws.iter_cols(), start=1):
                col_index = col_num - 1  # Convert to zero-based index
                width = column_widths.get(col_index, default_width)
                ws.column_dimensions[get_column_letter(col_num)].width = width

            wb.save(file_path)

            QMessageBox.information(self.view, "Sucesso", f"Dados exportados com sucesso para: {file_path}")
            
            # Open the exported file
            if sys.platform.startswith("win"):  # Windows
                os.startfile(file_path)
            elif sys.platform.startswith("linux"):  # Linux
                subprocess.run(["xdg-open", file_path])
            
        except Exception as e:
            QMessageBox.critical(self.view, "Erro", f"Falha ao exportar dados: {str(e)}")


            print(f"Falha ao exportar dados: {str(e)}")
    
    def import_zip_to_db(self):
        """Baixa e importa os dados do ZIP do Portal da Transparência baseado no YYYYMM fornecido pelo usuário."""
        # Obter YYYYMM do usuário
        current_yyyymm = datetime.now().strftime("%Y%m")
        yyyymm, ok = QInputDialog.getText(self.view, "Informe o Período", "Digite YYYYMM:", text=current_yyyymm)

        if not ok or not yyyymm.isdigit() or len(yyyymm) != 6:
            QMessageBox.warning(self.view, "Erro", "Entrada inválida! O formato deve ser YYYYMM.")
            return

        # Construir URL do arquivo ZIP
        zip_url = f"https://portaldatransparencia.gov.br/download-de-dados/cpgf/{yyyymm}"

        try:
            # Baixar o arquivo ZIP
            response = requests.get(zip_url, stream=True)
            response.raise_for_status()
            zip_data = BytesIO(response.content)
            
            with zipfile.ZipFile(zip_data, "r") as zip_ref:
                csv_files = [f for f in zip_ref.namelist() if f.endswith(".csv")]
                if not csv_files:
                    QMessageBox.warning(self.view, "Erro", "Nenhum arquivo CSV encontrado no ZIP.")
                    return

                csv_filename = csv_files[0]  # Pegamos o primeiro arquivo CSV encontrado
                with zip_ref.open(csv_filename) as csv_file:
                    csv_data = BytesIO(csv_file.read())  # Lê o arquivo extraído para um buffer de memória
                    csv_data.seek(0)
                    result = chardet.detect(csv_data.read(100000))
                    encoding_detected = result["encoding"] if result["encoding"] else "utf-8"
                    csv_data.seek(0)
                    df = pd.read_csv(csv_data, dtype=str, sep=None, engine="python", encoding=encoding_detected)

            # Mapeamento de colunas
            column_mapping = {
                "CÓDIGO ÓRGÃO SUPERIOR": "cod_orgao_superior",
                "NOME ÓRGÃO SUPERIOR": "nome_orgao_superior",
                "CÓDIGO ÓRGÃO": "cod_orgao",
                "NOME ÓRGÃO": "nome_orgao",
                "CÓDIGO UNIDADE GESTORA": "cod_unidade_gestora",
                "NOME UNIDADE GESTORA": "nome_unidade_gestora",
                "ANO EXTRATO": "ano_extrato",
                "MÊS EXTRATO": "mes_extrato",
                "CPF PORTADOR": "cpf_portador",
                "NOME PORTADOR": "nome_portador",
                "CNPJ OU CPF FAVORECIDO": "cnpj_cpf_favorecido",
                "NOME FAVORECIDO": "nome_favorecido",
                "TRANSAÇÃO": "transacao",
                "DATA TRANSAÇÃO": "data_transacao",
                "VALOR TRANSAÇÃO": "valor_transacao",
            }
            df.rename(columns=column_mapping, inplace=True)

            # Filtrar apenas os códigos de órgão desejados
            df = df[df["cod_orgao"].isin(["52131", "52132"])]

            if df.empty:
                QMessageBox.warning(self.view, "Aviso", "Nenhum dado relevante encontrado para os órgãos 52131 e 52132.")
                return

            # Corrige valores monetários e converte para float
            if "valor_transacao" in df.columns:
                df["valor_transacao"] = (
                    df["valor_transacao"].str.replace(",", ".", regex=True)
                    .str.replace("[^0-9.]", "", regex=True)
                    .astype(float)
                )

            # Inserir dados no banco de dados
            with self.model.database_manager as conn:
                df.to_sql("tabela_cartao_corporativo", conn, if_exists="append", index=False)

            QMessageBox.information(self.view, "Sucesso", "Dados importados com sucesso!")
            self.view.model.select()

        except requests.exceptions.RequestException as e:
            QMessageBox.critical(self.view, "Erro", f"Falha ao baixar o arquivo ZIP: {e}")
        except Exception as e:
            QMessageBox.critical(self.view, "Erro", f"Falha ao importar os dados: {e}")

            
    def row_double_clicked(self, row_data):
        """Handles row double-click event."""
        pass

